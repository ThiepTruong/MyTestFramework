#!/usr/bin/python
from openpyxl import load_workbook
import logging
import MyLogger
logger = logging.getLogger('ExcelUtility')

def load_excel(file_name):
    try:
        return load_workbook(file_name)
    except Exception as e:
        logger.debug(e.errno)
        logger.debug(e.strerror)
        
def get_sheet_data(my_workbook, sheet_name):
    return my_workbook[sheet_name]

def load_object_repo (my_workbook, my_sheet_name):
    sheet_data = get_sheet_data(my_workbook, my_sheet_name)
    row_num = len(sheet_data['A']) + 1
    col_num = len(sheet_data[1]) + 1
    object_repo = {}
    for row in sheet_data.iter_rows(min_row = 1, max_col = col_num, max_row = row_num):
        object_repo[row[0].value] = {'value': row[1].value, 'comment': ''}
    return object_repo

def load_test_data (my_workbook, my_sheet_name):
    sheet_data = get_sheet_data(my_workbook, my_sheet_name)
    row_num = len(sheet_data['A']) + 1
    col_num = len(sheet_data[1]) + 1
    test_data = {}
    
    for i in range(1, col_num):
        test_data[sheet_data.cell(row = 1, column = i).value] = []
        for j in range( 2, row_num):
            test_data[sheet_data.cell(row = 1, column = i).value].append(sheet_data.cell(row = j, column = i).value)
    return test_data

if __name__ == "__main__":
    MyLogger.setup_logger('../../ExecutionResult/Log/Log_')
    file_name = "../../TestData/CommonData.xlsx"
    my_workbook = load_excel(file_name)
    print load_object_repo(my_workbook, 'NewBug')
    print load_test_data(my_workbook, 'TestData')
